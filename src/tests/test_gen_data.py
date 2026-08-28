"""Regression tests for the deterministic streaming data generator (D10)."""

import csv
import tempfile
import unittest
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from scripts.gen_data import FIELDS, generate_dataset


class TestGenerateData(unittest.TestCase):
    def test_csv_dan_xlsx_cocok_deterministik_dengan_duplikat_sengaja(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "first"
            second = root / "second"
            generate_dataset(rows=8, seed=42, out=first, duplicates=2)
            generate_dataset(rows=8, seed=42, out=second, duplicates=2)

            with (first / "records.csv").open(encoding="utf-8", newline="") as file:
                csv_rows = list(csv.reader(file))
            with (second / "records.csv").open(encoding="utf-8", newline="") as file:
                self.assertEqual(csv_rows, list(csv.reader(file)))

            workbook = load_workbook(first / "records.xlsx", read_only=True, data_only=True)
            worksheet = workbook["records"]
            xlsx_rows = [
                [str(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            workbook.close()

            self.assertEqual(list(FIELDS), csv_rows[0])
            self.assertEqual(csv_rows, xlsx_rows)
            self.assertEqual(9, len(csv_rows))
            refs = [row[0] for row in csv_rows[1:]]
            self.assertEqual({"REC-000001": 2, "REC-000002": 2}, {
                ref: count for ref, count in Counter(refs).items() if count > 1
            })

    def test_parameter_invalid_ditolak(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "duplicates"):
                generate_dataset(1, 42, Path(temp_dir), duplicates=1)

    def test_output_yang_ada_tidak_ditimpa(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            out = Path(temp_dir)
            csv_path = out / "records.csv"
            csv_path.write_text("raw-original\n", encoding="utf-8")

            with self.assertRaisesRegex(FileExistsError, "tidak ditimpa"):
                generate_dataset(8, 42, out, duplicates=2)

            self.assertEqual("raw-original\n", csv_path.read_text(encoding="utf-8"))
            self.assertFalse((out / "records.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
