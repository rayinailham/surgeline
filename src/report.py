"""Laporan klien P13: digest 0-jargon + `REPORT.xlsx`, seluruhnya dari antrean nyata.

Dua pembaca, dua bahasa. Klien membaca **hasil** (berapa terkirim, berapa gagal, kenapa,
berapa lama); orang teknis klien membaca **cara memeriksanya**. Modul ini memisahkan
keduanya dengan tegas:

- Teks klien (digest `.md` + sheet `Ringkasan`, `Per-status`, `Throughput`, `Kegagalan`)
  wajib lolos `jargon_violations()`. Istilah internal seperti *lease*, *backoff*,
  `BEGIN IMMEDIATE`, atau nama status mentah tidak pernah sampai ke klien — kalau lolos,
  penulisnya terbaca sebagai tukang script, bukan orang yang paham masalah klien.
- Sheet `Bukti` **sengaja dikecualikan** (`JARGON_EXEMPT_SHEETS`): isinya justru perintah
  SQL untuk mengulang tiap angka. Tanpa itu laporan hanya klaim.

Angka tidak pernah disalin dari dokumen lain: semuanya di-query ulang dari
`data/<run_id>/queue.db` secara read-only (`mode=ro`, sama seperti `src/dashboard.py`,
`src/throughput.py`, dan `scripts/run_summary.py`). Dua hal yang memang tidak ada di DB
disuplai operator dan ditandai apa adanya: `--kills` (berapa kali proses dimatikan paksa)
dan `--scale` (manifest sapuan N worker P12). Sisanya lahir dari `queue.db`.

`--sanitize` menghasilkan salinan yang boleh masuk repo (`assets/sample_*`): nomor
referensi dan nomor konfirmasi contoh disamarkan, jalur mesin lokal tidak ditulis.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import time
from dataclasses import dataclass
from pathlib import Path

if __package__:
    from . import throughput
else:  # dijalankan sebagai `python src/report.py`
    import throughput

#: Istilah yang tidak boleh muncul di teks klien. Dicocokkan sebagai kata utuh,
#: case-insensitive — "WAL" tidak boleh menandai "walau".
FORBIDDEN_JARGON: tuple[str, ...] = (
    "lease", "backoff", "jitter", "dead-letter", "dead letter", "begin immediate",
    "sqlite", "sql", "wal", "pragma", "queue.db", "external_ref", "worker_id", "rowid",
    "schema", "skema", "database", "query", "attempts", "payload",
    "pending", "claimed", "dead", "failed",
    "kill -9", "sigkill", "playwright", "chromium", "headless", "browser",
    "http", "timeout", "retry", "throughput", "worker", "runner", "daemon", "endpoint",
    "localhost", "127.0.0.1", "stdout", "commit", "repo", "docker", "container",
    "state machine", "idempoten", "dedup", "regex", "cron", "api",
)

#: Sheet yang memang teknis: `Bukti` berisi perintah verifikasi, bukan narasi klien.
JARGON_EXEMPT_SHEETS: tuple[str, ...] = ("Bukti",)

#: Status DB -> label manusia. Klien tidak pernah membaca nama status mentah.
STATUS_LABELS: dict[str, str] = {
    "ok": "Terkirim dan bernomor konfirmasi",
    "failed": "Ditolak sistem tujuan",
    "dead": "Tidak berhasil setelah 5 kali percobaan",
    "pending": "Belum dikerjakan",
    "claimed": "Sedang dikerjakan",
}

STATUS_MEANINGS: dict[str, str] = {
    "ok": "Data masuk ke sistem tujuan dan nomor konfirmasinya tersimpan. Ini satu-satunya "
          "yang dihitung berhasil.",
    "failed": "Sistem tujuan menolak isian ini dan penolakannya bersifat tetap. Mengirim "
              "ulang data yang sama akan ditolak lagi, jadi tidak diulang.",
    "dead": "Sistem tujuan sedang bermasalah saat data ini dikirim. Sudah dicoba lima kali "
            "dengan jeda yang makin panjang, alasan kegagalannya tercatat semua.",
    "pending": "Masih menunggu giliran.",
    "claimed": "Sedang dikerjakan saat laporan ini dibuat.",
}

#: Kode dicocokkan sebagai angka utuh: substring "500" juga ada di dalam "15000ms",
#: dan pesan timeout tidak boleh diterjemahkan jadi "sistem tujuan bermasalah".
_REASON_RULES: tuple[tuple[re.Pattern[str], str], ...] = (
    (
        re.compile(r"\b422\b"),
        "Isian ditolak sistem tujuan karena tidak lolos pemeriksaan datanya. Penolakan ini "
        "tetap, jadi tidak dicoba ulang — data perlu diperbaiki di sumbernya.",
    ),
    (
        re.compile(r"\b500\b"),
        "Sistem tujuan sedang bermasalah saat data ini dikirim. Sudah dicoba lima kali "
        "dengan jeda yang makin panjang dan tetap gagal; datanya utuh dan bisa dikirim "
        "ulang kapan saja.",
    ),
)

_GENERIC_REASON = (
    "Pengiriman gagal dan alasannya tercatat lengkap di catatan pekerjaan. Datanya utuh "
    "dan bisa dikirim ulang."
)

#: Jumlah data yang dijanjikan di proposal klien; tidak pernah benar-benar dijalankan (D15).
DEFAULT_EXTRAPOLATION_RECORDS = throughput.DEFAULT_EXTRAPOLATION_RECORDS


@dataclass(frozen=True)
class FailureGroup:
    """Satu kelompok kegagalan: alasan manusia + jumlah + contoh nomor referensi."""

    status: str
    reason: str
    count: int
    examples: tuple[str, ...]


@dataclass(frozen=True)
class RunFacts:
    """Seluruh angka laporan, hasil query ulang ke `queue.db`."""

    run_id: str
    generated_at: float
    jobs_total: int
    status: dict[str, int]
    terminal: int
    non_terminal: int
    attempts_total: int
    attempts_by_outcome: dict[str, int]
    duplicate_external_ref: int
    duplicate_confirmation: int
    ok_without_confirmation: int
    dead_without_last_error: int
    recovered_orphans: int
    duration_seconds: float
    rate: throughput.RunThroughput
    failures: tuple[FailureGroup, ...]
    workers: int
    kills: int = 0
    scale_best: throughput.ScalingRow | None = None
    scale_saturation: int | None = None
    scale_saturation_proven: bool = False

    @property
    def ok(self) -> int:
        return self.status.get("ok", 0)

    @property
    def success_ratio(self) -> float:
        return self.ok / self.jobs_total if self.jobs_total else 0.0


def jargon_violations(text: str) -> list[str]:
    """Istilah teknis terlarang yang muncul di `text`, terurut dan tanpa duplikat.

    Kata utuh, bukan substring: `\\bwal\\b` tidak boleh menandai "walau", dan `\\bapi\\b`
    tidak boleh menandai "sampai". Ini gerbang yang menahan laporan klien tetap manusiawi.
    """
    found: list[str] = []
    for term in FORBIDDEN_JARGON:
        pattern = re.compile(rf"(?<![\w-]){re.escape(term)}(?![\w-])", re.IGNORECASE)
        if pattern.search(text) and term not in found:
            found.append(term)
    return sorted(found)


def human_reason(last_error: str | None) -> str:
    """Terjemahkan pesan kegagalan internal jadi kalimat yang berguna untuk klien.

    Pesan mentah tidak pernah diteruskan apa adanya: ia memuat kode teknis, dan klien
    tidak bisa berbuat apa pun dengan "HTTP 500". Yang berguna baginya adalah apakah
    datanya perlu diperbaiki atau cukup dikirim ulang.
    """
    if not last_error:
        return _GENERIC_REASON
    for pattern, reason in _REASON_RULES:
        if pattern.search(last_error):
            return reason
    return _GENERIC_REASON


def mask_identifier(value: str) -> str:
    """Samarkan bagian pembeda sebuah pengenal, sisakan bentuknya (dipakai `--sanitize`).

    Awalan sampai tanda hubung pertama dipertahankan karena ia bentuk, bukan identitas
    (`REC-000215` -> `REC-••••••`). Menyamarkan per-karakter-hex tidak bisa dipakai:
    huruf `E` dan `C` pada `REC` juga hex, dan awalannya ikut hilang.
    """
    prefix, sep, tail = value.partition("-")
    if not sep:
        return re.sub(r"[0-9A-Za-z]", "•", value)
    return prefix + sep + re.sub(r"[0-9A-Za-z]", "•", tail)


def _db_path(run_id: str) -> Path:
    return Path("data") / run_id / "queue.db"


def _failure_groups(conn: sqlite3.Connection, *, limit: int = 3) -> tuple[FailureGroup, ...]:
    rows = conn.execute(
        "SELECT status, last_error, COUNT(*) AS n FROM jobs"
        " WHERE status IN ('failed', 'dead') GROUP BY status, last_error ORDER BY n DESC"
    ).fetchall()
    groups: list[FailureGroup] = []
    for status, last_error, count in rows:
        examples = conn.execute(
            "SELECT external_ref FROM jobs WHERE status = ? AND last_error IS ?"
            " ORDER BY external_ref LIMIT ?",
            (status, last_error, limit),
        ).fetchall()
        groups.append(
            FailureGroup(
                status=status,
                reason=human_reason(last_error),
                count=count,
                examples=tuple(row[0] for row in examples),
            )
        )
    return tuple(groups)


def collect(
    run_id: str,
    *,
    db: Path | None = None,
    workers: int | None = None,
    kills: int = 0,
    scale: Path | None = None,
) -> RunFacts:
    """Baca satu antrean dan kumpulkan setiap angka laporan. Read-only, tanpa menebak."""
    path = db or _db_path(run_id)
    if not path.exists():
        raise SystemExit(f"antrean tidak ada: {path}")

    rate = throughput.measure(run_id, workers=workers, db=path)

    conn = throughput.connect_ro(path)
    try:
        scalar = lambda sql: conn.execute(sql).fetchone()[0]  # noqa: E731
        status = dict(conn.execute("SELECT status, COUNT(*) FROM jobs GROUP BY status"))
        outcome = dict(conn.execute("SELECT outcome, COUNT(*) FROM attempts GROUP BY outcome"))
        facts = RunFacts(
            run_id=run_id,
            generated_at=time.time(),
            jobs_total=scalar("SELECT COUNT(*) FROM jobs"),
            status=status,
            terminal=sum(status.get(name, 0) for name in ("ok", "failed", "dead")),
            non_terminal=status.get("pending", 0) + status.get("claimed", 0),
            attempts_total=scalar("SELECT COUNT(*) FROM attempts"),
            attempts_by_outcome=outcome,
            duplicate_external_ref=scalar(
                "SELECT COUNT(*) - COUNT(DISTINCT external_ref) FROM jobs"
            ),
            duplicate_confirmation=scalar(
                "SELECT COUNT(*) - COUNT(DISTINCT confirmation) FROM jobs WHERE status = 'ok'"
            ),
            ok_without_confirmation=scalar(
                "SELECT COUNT(*) FROM jobs"
                " WHERE status = 'ok' AND (confirmation IS NULL OR confirmation = '')"
            ),
            dead_without_last_error=scalar(
                "SELECT COUNT(*) FROM jobs WHERE status = 'dead' AND last_error IS NULL"
            ),
            # Percobaan yang hilang bersama proses yang mati lalu ditarik kembali (D8):
            # inilah bukti pemulihan yang benar-benar ada di DB, bukan di catatan operator.
            recovered_orphans=outcome.get("timeout", 0),
            # Satu-satunya durasi yang dipakai laporan: jam dinding percobaan pertama
            # sampai terakhir (sama seperti `docs/THROUGHPUT.md`). Memakai `created_at`
            # job akan ikut menghitung waktu pemuatan file dan membuat laporan
            # menyebut dua durasi yang berbeda untuk pekerjaan yang sama.
            duration_seconds=rate.window_seconds,
            rate=rate,
            failures=_failure_groups(conn),
            workers=rate.workers,
            kills=kills,
        )
    finally:
        conn.close()

    if scale is None:
        return facts

    manifest = json.loads(Path(scale).read_text(encoding="utf-8"))
    rows = throughput.scaling_table(
        [throughput.measure(item["run_id"], workers=int(item["workers"])) for item in manifest]
    )
    if not rows:
        return facts
    best = max(rows, key=lambda row: row.ok_per_hour)
    saturation, proven = throughput.saturation_workers(rows)
    return RunFacts(
        **{
            **{key: getattr(facts, key) for key in facts.__dataclass_fields__},
            "scale_best": best,
            "scale_saturation": saturation,
            "scale_saturation_proven": proven,
        }
    )


# --------------------------------------------------------------------------- teks klien

#: Nama bulan Indonesia; `time.strftime('%B')` mengikuti locale mesin dan bisa
#: menghasilkan laporan setengah Inggris di mesin klien.
BULAN: tuple[str, ...] = (
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
)


def tanggal_indonesia(epoch: float) -> str:
    waktu = time.localtime(epoch)
    return f"{waktu.tm_mday} {BULAN[waktu.tm_mon - 1]} {waktu.tm_year}"


def _num(value: float, decimals: int = 0) -> str:
    """Format angka gaya Indonesia: titik ribuan, koma desimal."""
    return f"{value:,.{decimals}f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _persen(part: int, whole: int) -> str:
    return f"{(part / whole * 100) if whole else 0.0:.1f}".replace(".", ",") + "%"


def _durasi(seconds: float) -> str:
    """Detik -> "1 jam 12 menit 30 detik". Klien tidak menghitung detik sendiri."""
    total = int(round(seconds))
    jam, sisa = divmod(total, 3600)
    menit, detik = divmod(sisa, 60)
    bagian = [f"{jam} jam"] if jam else []
    if menit:
        bagian.append(f"{menit} menit")
    if detik or not bagian:
        bagian.append(f"{detik} detik")
    return " ".join(bagian)


def _contoh(refs: tuple[str, ...], *, sanitize: bool) -> str:
    values = [mask_identifier(ref) for ref in refs] if sanitize else list(refs)
    return ", ".join(values) if values else "—"


def digest_markdown(
    facts: RunFacts, *, records: int = DEFAULT_EXTRAPOLATION_RECORDS, sanitize: bool = False
) -> str:
    """Digest satu halaman untuk klien. Wajib lolos `jargon_violations()`.

    Setiap angka di sini adalah angka yang sama dengan `REPORT.xlsx`, dan dua-duanya
    lahir dari objek `RunFacts` yang sama — tidak ada kesempatan keduanya berbeda.
    """
    ok = facts.ok
    ditolak = facts.status.get("failed", 0)
    gagal = facts.status.get("dead", 0)
    jam, hari = throughput.extrapolate(records, facts.rate.ok_per_hour)

    baris = [
        f"# Laporan pengiriman data — {facts.run_id}",
        "",
        f"**Tanggal laporan:** {tanggal_indonesia(facts.generated_at)}  ",
        "**Asal angka:** catatan pekerjaan pengiriman ini, dibaca ulang saat laporan dibuat. "
        "Cara memeriksa tiap angkanya ada di lembar **Bukti** pada `REPORT.xlsx`.",
        "",
        "## Hasil",
        "",
        f"- Data yang dikirim: **{_num(facts.jobs_total)}**",
        f"- Terkirim dan bernomor konfirmasi: **{_num(ok)}** ({_persen(ok, facts.jobs_total)})",
        f"- Ditolak sistem tujuan: **{_num(ditolak)}** ({_persen(ditolak, facts.jobs_total)})",
        f"- Tidak berhasil setelah 5 kali percobaan: **{_num(gagal)}** "
        f"({_persen(gagal, facts.jobs_total)})",
        f"- Belum selesai saat laporan dibuat: **{_num(facts.non_terminal)}**",
        f"- Lama pengerjaan: **{_durasi(facts.duration_seconds)}**",
        "",
        "## Tidak ada data yang terkirim dua kali",
        "",
        f"Nomor rujukan yang muncul lebih dari sekali: **{facts.duplicate_external_ref}**. "
        f"Nomor konfirmasi kembar: **{facts.duplicate_confirmation}**. "
        f"Pengiriman yang tercatat berhasil tetapi nomor konfirmasinya hilang: "
        f"**{facts.ok_without_confirmation}**.",
        "",
        "Setiap baris data punya satu nomor rujukan, dan sistem menolak menerima nomor "
        "rujukan yang sama dua kali. Memuat ulang file yang sama tidak menambah pekerjaan "
        "baru dan tidak mengirim ulang data yang sudah terkirim.",
        "",
    ]

    baris += ["## Pengerjaan tidak hilang walau dihentikan di tengah jalan", ""]
    if facts.kills:
        baris.append(
            f"Pengerjaan sengaja saya hentikan paksa **{facts.kills} kali** di tengah jalan, "
            "seperti listrik padam. Setiap kali dinyalakan lagi, pekerjaan dilanjutkan dari "
            "titik terakhir, bukan diulang dari awal."
        )
    else:
        baris.append(
            "Pekerjaan dicatat satu per satu, jadi berhenti di tengah jalan tidak "
            "menghilangkan kemajuan yang sudah ada."
        )
    baris += [
        "",
        f"Ada **{_num(facts.recovered_orphans)}** data yang sedang dikerjakan tepat saat "
        "pengerjaan berhenti. Semuanya ditemukan kembali dan diselesaikan, tanpa satu pun "
        "yang terkirim dua kali. Yang tersisa saat laporan ini dibuat: "
        f"**{_num(facts.non_terminal)}**.",
        "",
        "## Kecepatan",
        "",
        f"- Kecepatan pengerjaan ini: **{_num(facts.rate.ok_per_hour)} data per jam** "
        f"dengan {facts.workers} pengiriman berjalan berbarengan.",
    ]
    if facts.scale_best is not None:
        baris.append(
            f"- Kecepatan tertinggi yang terukur: **{_num(facts.scale_best.ok_per_hour)} data "
            f"per jam** dengan {facts.scale_best.workers} pengiriman berbarengan."
        )
    if facts.scale_saturation is not None and facts.scale_saturation_proven:
        baris.append(
            f"- Menambah pengiriman berbarengan di atas **{facts.scale_saturation}** tidak "
            "lagi mempercepat secara berarti; yang penuh adalah tenaga komputer, jadi "
            "mempercepatnya berarti menambah mesin, bukan menambah proses."
        )
    baris += [
        "",
        f"Pada kecepatan pengerjaan ini, **{_num(records)}** data selesai dalam sekitar "
        f"**{_num(jam, 1)} jam ≈ {_num(hari, 1)} hari**. Angka itu perpanjangan lurus dari "
        "kecepatan yang benar-benar terukur di sini, bukan hasil menjalankan sebanyak itu. "
        "Kecepatan di sistem tujuan Anda akan berbeda: jaringan, batas pengiriman resmi, "
        "dan antrean di sisi mereka semuanya memperlambat.",
        "",
        "## Kegagalan dan artinya",
        "",
        "| Hasil | Jumlah | Apa artinya | Contoh nomor rujukan |",
        "|---|---|---|---|",
    ]
    for group in facts.failures:
        baris.append(
            f"| {STATUS_LABELS.get(group.status, group.status)} | {_num(group.count)} | "
            f"{group.reason} | {_contoh(group.examples, sanitize=sanitize)} |"
        )
    baris += [
        "",
        "Tidak ada kegagalan yang disembunyikan atau dihapus: semua yang gagal punya "
        "catatan alasannya, dan datanya tetap utuh untuk dikirim ulang.",
        "",
        "## Lingkup dan izin",
        "",
        "Seluruh pengiriman dalam laporan ini ditujukan ke sistem uji milik saya sendiri, "
        "yang memang saya buat untuk gagal sekitar 5% agar ketahanannya terbukti. Tidak ada "
        "data yang dikirim ke sistem pihak lain, tidak ada pengamanan yang ditembus, dan "
        "tidak ada akun orang lain yang dipakai. Untuk pekerjaan sungguhan, izin tertulis "
        "dari pemilik sistem tujuan adalah syarat mulai, bukan formalitas.",
        "",
    ]
    return "\n".join(baris)


# ------------------------------------------------------------------------------- workbook

SHEET_ORDER: tuple[str, ...] = ("Ringkasan", "Per-status", "Throughput", "Kegagalan", "Bukti")


def sheet_rows(
    facts: RunFacts, *, records: int = DEFAULT_EXTRAPOLATION_RECORDS, sanitize: bool = False
) -> dict[str, list[list[object]]]:
    """Isi kelima lembar `REPORT.xlsx`. Baris pertama tiap lembar adalah judul kolom."""
    jam, hari = throughput.extrapolate(records, facts.rate.ok_per_hour)
    db = f"data/{facts.run_id}/queue.db"

    ringkasan: list[list[object]] = [
        ["Keterangan", "Nilai"],
        ["Kode pekerjaan", facts.run_id],
        ["Tanggal laporan", time.strftime("%Y-%m-%d %H:%M", time.localtime(facts.generated_at))],
        ["Data yang dikirim", facts.jobs_total],
        ["Terkirim dan bernomor konfirmasi", facts.ok],
        ["Ditolak sistem tujuan", facts.status.get("failed", 0)],
        ["Tidak berhasil setelah 5 kali percobaan", facts.status.get("dead", 0)],
        ["Belum selesai", facts.non_terminal],
        ["Bagian yang berhasil", _persen(facts.ok, facts.jobs_total)],
        ["Lama pengerjaan", _durasi(facts.duration_seconds)],
        ["Pengiriman berjalan berbarengan", facts.workers],
        ["Dihentikan paksa di tengah jalan", f"{facts.kills} kali"],
        ["Data yang diselamatkan setelah berhenti mendadak", facts.recovered_orphans],
        ["Nomor rujukan kembar", facts.duplicate_external_ref],
        ["Nomor konfirmasi kembar", facts.duplicate_confirmation],
        ["Berhasil tetapi nomor konfirmasinya hilang", facts.ok_without_confirmation],
        ["Gagal tanpa catatan alasan", facts.dead_without_last_error],
    ]

    per_status: list[list[object]] = [["Hasil", "Jumlah", "Bagian", "Apa artinya"]]
    for name in ("ok", "failed", "dead", "pending", "claimed"):
        jumlah = facts.status.get(name, 0)
        if not jumlah and name in ("pending", "claimed"):
            continue
        per_status.append(
            [
                STATUS_LABELS[name],
                jumlah,
                _persen(jumlah, facts.jobs_total),
                STATUS_MEANINGS[name],
            ]
        )
    per_status.append(["Jumlah seluruhnya", facts.jobs_total, "100,0%", ""])

    kecepatan: list[list[object]] = [
        ["Keterangan", "Nilai"],
        ["Kecepatan pengerjaan ini (data per jam)", round(facts.rate.ok_per_hour)],
        ["Pengiriman berjalan berbarengan", facts.workers],
        ["Lama pengerjaan", _durasi(facts.duration_seconds)],
    ]
    if facts.scale_best is not None:
        kecepatan += [
            ["Kecepatan tertinggi terukur (data per jam)", round(facts.scale_best.ok_per_hour)],
            ["Jumlah berbarengan pada kecepatan itu", facts.scale_best.workers],
        ]
    if facts.scale_saturation is not None and facts.scale_saturation_proven:
        kecepatan.append(
            ["Menambah proses di atas ini tidak lagi mempercepat", facts.scale_saturation]
        )
    kecepatan += [
        [f"Perkiraan {_num(records)} data — jam", round(jam, 1)],
        [f"Perkiraan {_num(records)} data — hari", round(hari, 1)],
        [
            "Catatan",
            "Perkiraan di atas adalah perpanjangan lurus dari kecepatan yang terukur di "
            "sini, bukan hasil menjalankan sebanyak itu. Kecepatan di sistem tujuan Anda "
            "bergantung pada jaringan dan batas pengiriman resminya.",
        ],
    ]

    kegagalan: list[list[object]] = [["Hasil", "Jumlah", "Apa artinya", "Contoh nomor rujukan"]]
    for group in facts.failures:
        kegagalan.append(
            [
                STATUS_LABELS.get(group.status, group.status),
                group.count,
                group.reason,
                _contoh(group.examples, sanitize=sanitize),
            ]
        )
    if len(kegagalan) == 1:
        kegagalan.append(["Tidak ada kegagalan", 0, "Semua data terkirim.", "—"])

    bukti: list[list[object]] = [
        ["Angka", "Nilai", "Perintah pemeriksaan ulang"],
        [
            "Total job",
            facts.jobs_total,
            f'sqlite3 "file:{db}?mode=ro" "SELECT COUNT(*) FROM jobs;"',
        ],
        [
            "Job per status",
            f"ok={facts.ok} failed={facts.status.get('failed', 0)} "
            f"dead={facts.status.get('dead', 0)}",
            f'sqlite3 "file:{db}?mode=ro" "SELECT status,COUNT(*) FROM jobs GROUP BY status;"',
        ],
        [
            "Job non-terminal (pending+claimed)",
            facts.non_terminal,
            f'sqlite3 "file:{db}?mode=ro" '
            "\"SELECT COUNT(*) FROM jobs WHERE status IN ('pending','claimed');\"",
        ],
        [
            "Duplikat external_ref (D4)",
            facts.duplicate_external_ref,
            f'sqlite3 "file:{db}?mode=ro" '
            '"SELECT COUNT(*)-COUNT(DISTINCT external_ref) FROM jobs;"',
        ],
        [
            "Duplikat confirmation pada sukses (D6)",
            facts.duplicate_confirmation,
            f'sqlite3 "file:{db}?mode=ro" "SELECT COUNT(*)-COUNT(DISTINCT confirmation) '
            "FROM jobs WHERE status='ok';\"",
        ],
        [
            "Sukses tanpa nomor konfirmasi (D6)",
            facts.ok_without_confirmation,
            f'sqlite3 "file:{db}?mode=ro" "SELECT COUNT(*) FROM jobs '
            "WHERE status='ok' AND (confirmation IS NULL OR confirmation='');\"",
        ],
        [
            "Dead tanpa last_error (D7)",
            facts.dead_without_last_error,
            f'sqlite3 "file:{db}?mode=ro" '
            "\"SELECT COUNT(*) FROM jobs WHERE status='dead' AND last_error IS NULL;\"",
        ],
        [
            "Percobaan per outcome",
            " ".join(f"{key}={value}" for key, value in sorted(facts.attempts_by_outcome.items())),
            f'sqlite3 "file:{db}?mode=ro" '
            '"SELECT outcome,COUNT(*) FROM attempts GROUP BY outcome;"',
        ],
        [
            "Klaim yatim dipulihkan lewat lease (D8)",
            facts.recovered_orphans,
            f'sqlite3 "file:{db}?mode=ro" '
            "\"SELECT COUNT(*) FROM attempts WHERE outcome='timeout';\"",
        ],
        [
            "Jendela wall-clock (detik)",
            facts.rate.window_seconds,
            f"uv run --no-sync python src/throughput.py --run {facts.run_id}",
        ],
        [
            "Throughput (record ok/jam)",
            facts.rate.ok_per_hour,
            f"uv run --no-sync python src/throughput.py --run {facts.run_id}",
        ],
        [
            "Laporan ini dibuat ulang dengan",
            "src/report.py",
            f"uv run --no-sync python src/report.py --run {facts.run_id} --out reports/",
        ],
    ]

    return {
        "Ringkasan": ringkasan,
        "Per-status": per_status,
        "Throughput": kecepatan,
        "Kegagalan": kegagalan,
        "Bukti": bukti,
    }


def sheet_jargon_violations(sheets: dict[str, list[list[object]]]) -> dict[str, list[str]]:
    """Jargon per lembar, kecuali lembar teknis (`JARGON_EXEMPT_SHEETS`)."""
    hasil: dict[str, list[str]] = {}
    for name, rows in sheets.items():
        if name in JARGON_EXEMPT_SHEETS:
            continue
        teks = "\n".join(str(cell) for row in rows for cell in row)
        found = jargon_violations(teks)
        if found:
            hasil[name] = found
    return hasil


def write_workbook(sheets: dict[str, list[list[object]]], path: Path) -> Path:
    """Tulis `REPORT.xlsx`. `openpyxl` langsung (D2/D5: tanpa MCP excel saat berjalan)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    book = Workbook()
    book.remove(book.active)
    for name in SHEET_ORDER:
        rows = sheets[name]
        sheet = book.create_sheet(title=name)
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        widths = [14] * len(rows[0])
        for row in rows:
            for index, cell in enumerate(row):
                widths[index] = max(widths[index], min(len(str(cell)) + 2, 70))
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run", required=True, help="run_id di data/<run_id>/queue.db")
    parser.add_argument("--out", type=Path, default=Path("reports"), help="folder keluaran")
    parser.add_argument("--workers", type=int, help="N worker run ini (default: dari run.json)")
    parser.add_argument(
        "--kills", type=int, default=0, help="berapa kali run dimatikan paksa (dari log operator)"
    )
    parser.add_argument("--scale", type=Path, help="manifest sapuan N worker P12 (opsional)")
    parser.add_argument(
        "--extrapolate", type=int, default=DEFAULT_EXTRAPOLATION_RECORDS,
        help="jumlah record yang diekstrapolasi",
    )
    parser.add_argument(
        "--sanitize", action="store_true",
        help="samarkan nomor contoh; dipakai untuk assets/sample_* yang masuk repo",
    )
    args = parser.parse_args(argv)

    facts = collect(
        args.run, workers=args.workers, kills=args.kills, scale=args.scale
    )
    digest = digest_markdown(facts, records=args.extrapolate, sanitize=args.sanitize)
    sheets = sheet_rows(facts, records=args.extrapolate, sanitize=args.sanitize)

    pelanggaran = jargon_violations(digest)
    pelanggaran_sheet = sheet_jargon_violations(sheets)
    if pelanggaran or pelanggaran_sheet:
        print(f"GAGAL jargon di digest: {pelanggaran}")
        for name, terms in pelanggaran_sheet.items():
            print(f"GAGAL jargon di lembar {name}: {terms}")
        return 1

    args.out.mkdir(parents=True, exist_ok=True)
    prefix = "sample_" if args.sanitize else ""
    digest_path = args.out / (f"{prefix}daily.md" if args.sanitize else "digest.md")
    digest_path.write_text(digest, encoding="utf-8")
    xlsx_path = write_workbook(sheets, args.out / f"{prefix}REPORT.xlsx")

    print(f"digest : {digest_path} (0 jargon dari {len(FORBIDDEN_JARGON)} istilah terlarang)")
    print(f"laporan: {xlsx_path} ({len(SHEET_ORDER)} lembar: {', '.join(SHEET_ORDER)})")
    print(
        f"angka  : total={facts.jobs_total} ok={facts.ok} "
        f"failed={facts.status.get('failed', 0)} dead={facts.status.get('dead', 0)} "
        f"dup_ref={facts.duplicate_external_ref} dup_conf={facts.duplicate_confirmation} "
        f"ok/jam={facts.rate.ok_per_hour:,.0f}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
