"""Laporan klien P13: gerbang 0-jargon, terjemahan kegagalan, dan angka dari DB.

Yang diuji bukan "file terbentuk" melainkan tiga hal yang bisa merusak laporan tanpa
kelihatan: istilah internal bocor ke teks klien, pesan kegagalan mentah diteruskan
apa adanya, dan angka laporan menyimpang dari isi antrean.
"""

from __future__ import annotations

import json
import sqlite3
import tempfile
import unittest
from pathlib import Path

from src import report
from src.schema import init_db

NOW = 1_000_000.0


class JargonTestCase(unittest.TestCase):
    """Gerbang bahasa: kata utuh, bukan substring."""

    def test_istilah_terlarang_terdeteksi(self) -> None:
        found = report.jargon_violations("Job dikembalikan setelah lease timeout habis.")
        self.assertIn("lease", found)
        self.assertIn("timeout", found)

    def test_kata_indonesia_tidak_kena_substring(self) -> None:
        # "walau" memuat "wal", "rapi" memuat "api", "sampai" memuat "amp".
        self.assertEqual(report.jargon_violations("Walau rapi, sampai sekarang aman."), [])

    def test_tanda_hubung_tidak_menyelundupkan_jargon(self) -> None:
        self.assertEqual(report.jargon_violations("dead-letter"), ["dead-letter"])

    def test_teks_manusia_lolos(self) -> None:
        self.assertEqual(
            report.jargon_violations(
                "Data terkirim dan nomor konfirmasinya tersimpan. Tidak ada yang kembar."
            ),
            [],
        )


class ReasonTestCase(unittest.TestCase):
    """Pesan mentah tidak pernah sampai ke klien; yang sampai adalah tindakannya."""

    def test_penolakan_permanen_tidak_dicoba_ulang(self) -> None:
        reason = report.human_reason("HTTP 422: record ditolak oleh chaos target")
        self.assertIn("tidak dicoba ulang", reason)
        self.assertEqual(report.jargon_violations(reason), [])

    def test_kegagalan_sistem_tujuan_dicoba_lima_kali(self) -> None:
        reason = report.human_reason("habis 5 percobaan: HTTP 500 dari target")
        self.assertIn("lima kali", reason)
        self.assertEqual(report.jargon_violations(reason), [])

    def test_pesan_mentah_tidak_pernah_diteruskan(self) -> None:
        mentah = "PlaywrightTimeoutError: page.goto exceeded 15000ms"
        self.assertNotIn("Playwright", report.human_reason(mentah))
        self.assertEqual(report.human_reason(mentah), report.human_reason(None))

    def test_mask_identifier_menyamarkan_angka_dan_hex(self) -> None:
        self.assertEqual(report.mask_identifier("REC-000215"), "REC-••••••")
        self.assertEqual(report.mask_identifier("SL-c3d44db9"), "SL-••••••••")


class QueueFixture:
    """Antrean kecil berisi kasus yang dipakai tiga kelas uji di bawah.

    Sengaja bukan `TestCase`: kalau kelas uji lain mewarisi `TestCase` berisi test,
    test milik induknya ikut dijalankan ulang tiap kelas anak.
    """

    def setUp(self) -> None:
        self.directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.directory.cleanup)
        self.db = Path(self.directory.name) / "queue.db"
        self.conn = init_db(self.db)
        self.addCleanup(self.conn.close)

    def seed_job(
        self,
        ref: str,
        status: str,
        *,
        confirmation: str | None = None,
        last_error: str | None = None,
    ) -> None:
        self.conn.execute(
            "INSERT INTO jobs (external_ref, payload, status, confirmation, last_error,"
            " created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (ref, json.dumps({"external_ref": ref}), status, confirmation, last_error, NOW, NOW),
        )

    def seed_attempt(self, ref: str, started: float, ended: float, outcome: str) -> None:
        self.conn.execute(
            "INSERT INTO attempts (external_ref, worker_id, started_at, ended_at, outcome)"
            " VALUES (?, ?, ?, ?, ?)",
            (ref, "w-1", started, ended, outcome),
        )

    def seed_run(self) -> report.RunFacts:
        """3 sukses, 1 penolakan tetap, 1 habis percobaan, 1 klaim yatim dipulihkan."""
        for index in range(3):
            ref = f"REC-{index:06d}"
            self.seed_job(ref, "ok", confirmation=f"SL-{index:08x}")
            self.seed_attempt(ref, NOW, NOW + 1, "ok")
        self.seed_job("REC-000100", "failed", last_error="HTTP 422: record ditolak")
        self.seed_attempt("REC-000100", NOW, NOW + 1, "permanent")
        self.seed_job("REC-000200", "dead", last_error="habis 5 percobaan: HTTP 500 dari target")
        self.seed_attempt("REC-000200", NOW, NOW + 1, "retryable")
        self.seed_attempt("REC-000200", NOW + 1, NOW + 3600, "timeout")
        return report.collect("uji", db=self.db, workers=4, kills=2)


class FactsTestCase(QueueFixture, unittest.TestCase):
    """Setiap angka laporan harus datang dari `queue.db`, bukan dari argumen."""

    def test_angka_ringkasan_dari_db(self) -> None:
        facts = self.seed_run()
        self.assertEqual(facts.jobs_total, 5)
        self.assertEqual(facts.ok, 3)
        self.assertEqual(facts.status.get("failed"), 1)
        self.assertEqual(facts.status.get("dead"), 1)
        self.assertEqual(facts.non_terminal, 0)
        self.assertEqual(facts.duplicate_external_ref, 0)
        self.assertEqual(facts.duplicate_confirmation, 0)
        self.assertEqual(facts.ok_without_confirmation, 0)
        self.assertEqual(facts.dead_without_last_error, 0)

    def test_klaim_yatim_dihitung_dari_audit_bukan_dari_operator(self) -> None:
        facts = self.seed_run()
        self.assertEqual(facts.recovered_orphans, 1)
        self.assertEqual(facts.attempts_by_outcome["timeout"], 1)

    def test_durasi_memakai_jendela_percobaan(self) -> None:
        facts = self.seed_run()
        self.assertEqual(facts.duration_seconds, facts.rate.window_seconds)
        self.assertAlmostEqual(facts.duration_seconds, 3600.0, places=1)

    def test_kegagalan_dikelompokkan_dengan_alasan_manusia(self) -> None:
        facts = self.seed_run()
        alasan = {group.status: group.reason for group in facts.failures}
        self.assertIn("tidak dicoba ulang", alasan["failed"])
        self.assertIn("lima kali", alasan["dead"])
        self.assertEqual(sum(group.count for group in facts.failures), 2)

    def test_collect_tidak_bisa_menulis_antrean(self) -> None:
        self.seed_run()
        conn = report.throughput.connect_ro(self.db)
        self.addCleanup(conn.close)
        with self.assertRaises(sqlite3.OperationalError):
            conn.execute("UPDATE jobs SET status = 'ok'")

    def test_antrean_tidak_ada_ditolak(self) -> None:
        with self.assertRaises(SystemExit):
            report.collect("uji", db=Path(self.directory.name) / "tidak-ada.db")


class DigestTestCase(QueueFixture, unittest.TestCase):
    """Digest adalah teks yang dibaca klien; ia tidak boleh mengklaim lebih dari DB."""

    def test_digest_nol_jargon(self) -> None:
        digest = report.digest_markdown(self.seed_run())
        self.assertEqual(report.jargon_violations(digest), [])

    def test_digest_memuat_angka_utama(self) -> None:
        digest = report.digest_markdown(self.seed_run())
        for potongan in ("**5**", "**3**", "**2 kali**", "60,0%"):
            self.assertIn(potongan, digest)

    def test_tanpa_kill_tidak_mengklaim_pernah_dimatikan(self) -> None:
        for index in range(2):
            ref = f"REC-{index:06d}"
            self.seed_job(ref, "ok", confirmation=f"SL-{index:08x}")
            self.seed_attempt(ref, NOW, NOW + 1, "ok")
        digest = report.digest_markdown(report.collect("uji", db=self.db, workers=1))
        self.assertNotIn("hentikan paksa", digest)

    def test_sanitize_menyamarkan_contoh_nomor(self) -> None:
        facts = self.seed_run()
        self.assertIn("REC-000100", report.digest_markdown(facts))
        disanitasi = report.digest_markdown(facts, sanitize=True)
        self.assertNotIn("REC-000100", disanitasi)
        self.assertIn("REC-••••••", disanitasi)

    def test_angka_indonesia(self) -> None:
        self.assertEqual(report._num(49950), "49.950")
        self.assertEqual(report._num(97.9, 1), "97,9")
        self.assertEqual(report._num(1234.5, 1), "1.234,5")
        self.assertEqual(report._durasi(2836.3), "47 menit 16 detik")


class WorkbookTestCase(QueueFixture, unittest.TestCase):
    """`REPORT.xlsx`: lima lembar, teks klien bersih, lembar Bukti sengaja teknis."""

    def test_lima_lembar_sesuai_urutan(self) -> None:
        sheets = report.sheet_rows(self.seed_run())
        self.assertEqual(tuple(sheets), report.SHEET_ORDER)
        self.assertEqual(len(report.SHEET_ORDER), 5)

    def test_lembar_klien_nol_jargon(self) -> None:
        self.assertEqual(report.sheet_jargon_violations(report.sheet_rows(self.seed_run())), {})

    def test_lembar_bukti_memang_teknis_dan_dikecualikan(self) -> None:
        sheets = report.sheet_rows(self.seed_run())
        teks = "\n".join(str(cell) for row in sheets["Bukti"] for cell in row)
        self.assertNotEqual(report.jargon_violations(teks), [])
        self.assertIn("Bukti", report.JARGON_EXEMPT_SHEETS)

    def test_workbook_tersimpan_dan_terbaca_ulang(self) -> None:
        from openpyxl import load_workbook

        path = Path(self.directory.name) / "REPORT.xlsx"
        report.write_workbook(report.sheet_rows(self.seed_run()), path)
        book = load_workbook(path)
        self.assertEqual(tuple(book.sheetnames), report.SHEET_ORDER)
        self.assertEqual(book["Ringkasan"]["A1"].value, "Keterangan")
        self.assertEqual(book["Ringkasan"]["B4"].value, 5)

    def test_per_status_menjumlah_seluruh_job(self) -> None:
        sheets = report.sheet_rows(self.seed_run())
        baris = sheets["Per-status"][1:-1]
        self.assertEqual(sum(row[1] for row in baris), 5)
        self.assertEqual(sheets["Per-status"][-1][1], 5)


if __name__ == "__main__":
    unittest.main()
