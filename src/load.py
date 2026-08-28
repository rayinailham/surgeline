"""Stream CSV/XLSX records into a run-local SQLite queue."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sqlite3
import time
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

if __package__:
    from .schema import RECORD_FIELDS, init_db
else:
    from schema import RECORD_FIELDS, init_db


EXTERNAL_REF_RE = re.compile(r"[A-Za-z0-9._-]{1,64}\Z")
EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+\Z")
RUN_ID_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9._-]{0,63}\Z")
INSERT_JOB = """
    INSERT OR IGNORE INTO jobs
        (external_ref, payload, status, created_at, updated_at)
    VALUES (?, ?, 'pending', ?, ?)
"""


class LoadError(ValueError):
    """Input cannot be loaded without violating the record contract."""


@dataclass(frozen=True)
class LoadStats:
    read: int
    inserted: int

    @property
    def duplicates(self) -> int:
        return self.read - self.inserted


def _validate_header(header: Sequence[object]) -> None:
    if tuple(header) != RECORD_FIELDS:
        raise LoadError(
            f"header harus {','.join(RECORD_FIELDS)}; diterima {','.join(map(str, header))}"
        )


def _records(path: Path) -> Iterator[tuple[int, dict[str, object]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as source:
            reader = csv.DictReader(source)
            _validate_header(reader.fieldnames or ())
            for row_number, row in enumerate(reader, start=2):
                yield row_number, dict(row)
        return
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            if worksheet is None:
                raise LoadError("file XLSX tidak punya worksheet")
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration as exc:
                raise LoadError("file XLSX kosong") from exc
            _validate_header(header)
            for row_number, values in enumerate(rows, start=2):
                if len(values) != len(RECORD_FIELDS):
                    raise LoadError(f"baris {row_number}: jumlah kolom tidak valid")
                yield row_number, dict(zip(RECORD_FIELDS, values, strict=True))
        finally:
            workbook.close()
        return
    raise LoadError("format input harus .csv atau .xlsx")


def validate_record(record: Mapping[str, object], row_number: int) -> dict[str, object]:
    """Validate and normalize one record before INSERT OR IGNORE can hide errors."""
    values = dict(record)
    errors: list[str] = []

    for field in ("external_ref", "full_name", "email", "policy_no"):
        if not isinstance(values.get(field), str):
            errors.append(f"{field} harus teks")

    external_ref = values.get("external_ref")
    full_name = values.get("full_name")
    email = values.get("email")
    policy_no = values.get("policy_no")
    notes = values.get("notes")

    if isinstance(external_ref, str) and not EXTERNAL_REF_RE.fullmatch(external_ref):
        errors.append("external_ref tidak valid")
    if isinstance(full_name, str) and (not full_name.strip() or len(full_name) > 120):
        errors.append("full_name tidak valid")
    if isinstance(email, str) and not EMAIL_RE.fullmatch(email):
        errors.append("email tidak valid")
    if isinstance(policy_no, str) and (not policy_no.strip() or len(policy_no) > 40):
        errors.append("policy_no tidak valid")
    if notes is None:
        values["notes"] = ""
    elif not isinstance(notes, str) or len(notes) > 500:
        errors.append("notes tidak valid")

    amount_value = values.get("amount")
    try:
        if isinstance(amount_value, bool) or not isinstance(amount_value, (str, int, float)):
            raise ValueError
        amount = float(amount_value)
        if not math.isfinite(amount) or amount <= 0:
            raise ValueError
        values["amount"] = amount
    except (TypeError, ValueError):
        errors.append("amount harus angka > 0")

    if errors:
        raise LoadError(f"baris {row_number}: {'; '.join(errors)}")
    return {field: values[field] for field in RECORD_FIELDS}


def _insert_batch(conn: sqlite3.Connection, batch: list[tuple[str, str, float, float]]) -> int:
    conn.execute("BEGIN")
    try:
        inserted = conn.executemany(INSERT_JOB, batch).rowcount
        conn.execute("COMMIT")
        return inserted
    except Exception:
        conn.execute("ROLLBACK")
        raise


def load_records(input_path: str | Path, db_path: str | Path, batch_size: int = 500) -> LoadStats:
    """Load records in bounded batches; database uniqueness performs deduplication."""
    if batch_size < 1:
        raise ValueError("batch_size harus >= 1")

    source = Path(input_path)
    if not source.is_file():
        raise LoadError(f"input tidak ditemukan: {source}")

    conn = init_db(db_path)
    read = inserted = 0
    batch: list[tuple[str, str, float, float]] = []
    try:
        for row_number, raw_record in _records(source):
            record = validate_record(raw_record, row_number)
            now = time.time()
            batch.append(
                (
                    str(record["external_ref"]),
                    json.dumps(record, ensure_ascii=False, separators=(",", ":")),
                    now,
                    now,
                )
            )
            read += 1
            if len(batch) >= batch_size:
                inserted += _insert_batch(conn, batch)
                batch.clear()
        if batch:
            inserted += _insert_batch(conn, batch)
    except (LoadError, OSError, sqlite3.Error):
        raise
    finally:
        conn.close()
    return LoadStats(read=read, inserted=inserted)


def _run_dir(run_id: str) -> Path:
    if not RUN_ID_RE.fullmatch(run_id):
        raise LoadError("run id harus 1-64 karakter aman dan diawali alfanumerik")
    return Path("data") / run_id


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--run", required=True)
    parser.add_argument("--batch-size", type=int, default=500)
    args = parser.parse_args()

    try:
        stats = load_records(args.input, _run_dir(args.run) / "queue.db", args.batch_size)
    except (LoadError, ValueError) as exc:
        parser.error(str(exc))
    print(
        f"dibaca={stats.read} dimasukkan={stats.inserted} "
        f"duplikat_ditolak={stats.duplicates}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
